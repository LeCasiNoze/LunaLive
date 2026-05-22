"""Convert L2N visitors CSV -> clean Excel .xlsx with:
- bold header
- frozen first row
- autofilter on all columns
- heavy clickers (5+ clicks) row highlighted in light yellow
- column widths auto-fit
"""

import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src = sys.argv[1] if len(sys.argv) > 1 else "exports/L2N_visitors_2026-05-12_to_2026-05-21.csv"
dst = sys.argv[2] if len(sys.argv) > 2 else "exports/L2N_visitors_2026-05-12_to_2026-05-21.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "L2N visitors"

# Read CSV
with open(src, "r", encoding="utf-8") as f:
    rows = list(csv.reader(f))

header = rows[0]
data = rows[1:]

# Write header
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

for col_idx, col_name in enumerate(header, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Index of "clicks" column for heavy-clicker highlight
try:
    clicks_idx = header.index("clicks")
except ValueError:
    clicks_idx = -1

heavy_fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
heavy_font = Font(bold=True, color="8B5A00")

# Write data
for r_idx, row in enumerate(data, start=2):
    is_heavy = False
    if clicks_idx >= 0:
        try:
            is_heavy = int(row[clicks_idx]) >= 5
        except (ValueError, IndexError):
            pass
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = thin_border
        if is_heavy:
            cell.fill = heavy_fill
            cell.font = heavy_font

# Freeze top row + autofilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)}"

# Auto-fit column widths (cap at 50)
for c_idx, col_name in enumerate(header, start=1):
    max_len = len(col_name)
    for row in data:
        if c_idx - 1 < len(row):
            max_len = max(max_len, len(str(row[c_idx - 1])))
    ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 50)

# Row height for header
ws.row_dimensions[1].height = 38

wb.save(dst)
print(f"OK: {dst} ({len(data)} rows)")
